import openpyxl

wb = openpyxl.load_workbook("exel.xlsx")
hoja_otros = wb.create_sheet("olimpidas")
print(wb.sheetnames)
hoja = hoja_otros

datos = [("USA", 46, 12, 5), ("China", 38, 20, 7), ("UK", 29, 7, 7), ("Russia", 22, 10, 9), ("South Korea", 13, 3, 2),
         ("Germany", 11, 7, 4)]
for dato in datos:
    hoja_otros.append(dato)

cabecera = hoja_otros.append(["Pais", "Oros", "Platas", "Bronces"])
hoja_otros.move_range("A7:D7", rows=-6, cols=0)

wb.save("exel.xlsx")




